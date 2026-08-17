#!/usr/bin/env python3
"""Выгрузить тексты подсказок из Google-таблицы в hint_texts.json.

Таблица — первоисточник по озвучке ([[clc-voice-lines-spreadsheet]]): Эдуард
правит текст там, и только потом идёт переозвучка. Локальной копии на замках
нет, поэтому чат гейммастера на игровом пульте берёт текст из этого файла.

    python3 tools/export_hint_texts.py          # перезаписать hint_texts.json

Файл едет на замки обычной раскаткой. Правили тексты в таблице — прогнать
заново и раскатать; ничего в коде менять не нужно.

Что внутри файла:
    texts    id → {ru, en, ar, ar_kw, fr, uk, pl}
    speakers префикс id → имена персонажа по языкам (для строки «Борис: hint_32_d»)
    esp32    устройство → список id по порядку «Playing Hint N» в его прошивке
"""

import json
import os
import subprocess
import sys

SHEET = "1cd-mAtyg0jatVvJhhAMhLVfhgawgvu7X2SwUum6e-iw"
REPO = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUT = os.path.join(REPO, "hint_texts.json")

# Столбцы вкладки hints. Порядок ОТЛИЧАЕТСЯ от вкладки stories: там
# Polish идёт раньше Ukranian, здесь наоборот — проверено по шапке.
COLUMNS = ["ru", "en", "uk", "pl", "fr", "ar", "ar_kw"]

# Кто произносит подсказку — на всех языках квеста. Ключ: либо полный id, либо
# префикс до последнего подчёркивания. Группировка взята из прошивки главной
# платы (массивы *Hints[]) и из того, какому устройству принадлежит этап.
#
# Имена правятся ЗДЕСЬ — это единственное место, откуда их берёт пульт. Оператор
# видит имя на том языке, на котором идёт игра: русское имя рядом с французским
# текстом смотрелось нелепо.
#
# Проверено Настей 17.08.2026 по САМОЙ ОЗВУЧКЕ: имена вытащены из уже
# записанных реплик каждого языка, а не переведены на слух. Где локализация имя
# даёт — берём её, иначе пульт разойдётся с тем, что слышат игроки.
#
# Гном, Гоблин, Рыцарь и Кузнец в текстах не называются ВООБЩЕ, ни на одном
# языке, включая русский — для них авторитета нет, это наши подписи для
# оператора. Появится имя в озвучке — приоритет за ней.
#
# ⚠ Осталось показать носителю: арабское «الغوبلن» (гоблин) — транслитерация,
# в кувейтской локализации может быть привычнее другое слово.
SPEAKERS = {
    # Настя 17.08: арабское имя взято из его же реплики story_2a («اسمي سيغنس»),
    # украинское — «Сигнус» через «и», как записано в озвучке. По правилам
    # украинского было бы «Сігнус», но пульт должен совпадать со звуком.
    "dragon_crystal": {"ru": "Дракон Сигнус", "en": "Signus the Dragon", "fr": "Le Dragon Signus",
                       "uk": "Дракон Сигнус", "pl": "Smok Signus", "ar": "التنين سيغنس"},
    "hint_2":  {"ru": "Дракон Сигнус", "en": "Signus the Dragon", "fr": "Le Dragon Signus",
                "uk": "Дракон Сигнус", "pl": "Smok Signus", "ar": "التنين سيغنس"},
    "hint_5":  {"ru": "Дракон Сигнус", "en": "Signus the Dragon", "fr": "Le Dragon Signus",
                "uk": "Дракон Сигнус", "pl": "Smok Signus", "ar": "التنين سيغنس"},
    "hint_3":  {"ru": "Студент", "en": "Student", "fr": "Étudiant",
                "uk": "Студент", "pl": "Student", "ar": "الطالب"},
    "hint_37": {"ru": "Студент", "en": "Student", "fr": "Étudiant",
                "uk": "Студент", "pl": "Student", "ar": "الطالب"},
    "hint_38": {"ru": "Студент", "en": "Student", "fr": "Étudiant",
                "uk": "Студент", "pl": "Student", "ar": "الطالب"},
    "hint_44": {"ru": "Студент", "en": "Student", "fr": "Étudiant",
                "uk": "Студент", "pl": "Student", "ar": "الطالب"},
    "hint_6":  {"ru": "Профессор", "en": "Professor", "fr": "Professeur",
                "uk": "Професор", "pl": "Profesor", "ar": "الأستاذ"},
    "hint_10": {"ru": "Профессор", "en": "Professor", "fr": "Professeur",
                "uk": "Професор", "pl": "Profesor", "ar": "الأستاذ"},
    "hint_11": {"ru": "Профессор", "en": "Professor", "fr": "Professeur",
                "uk": "Професор", "pl": "Profesor", "ar": "الأستاذ"},
    "hint_14": {"ru": "Гном", "en": "Dwarf", "fr": "Nain",
                "uk": "Гном", "pl": "Krasnal", "ar": "القزم"},
    "hint_17": {"ru": "Ведьма", "en": "Witch", "fr": "Sorcière",
                "uk": "Відьма", "pl": "Wiedźma", "ar": "الساحرة"},
    "hint_19": {"ru": "Рыцарь", "en": "Knight", "fr": "Chevalier",
                "uk": "Лицар", "pl": "Rycerz", "ar": "الفارس"},
    "hint_23": {"ru": "Гоблин-банкир", "en": "Goblin banker", "fr": "Gobelin banquier",
                "uk": "Гоблін-банкір", "pl": "Goblin bankier", "ar": "الغوبلن المصرفي"},
    # Настя: не транслитерация, а слово из локализации — hint_16_c, «كهف الغول».
    "hint_26": {"ru": "Тролль", "en": "Troll", "fr": "Troll",
                "uk": "Троль", "pl": "Troll", "ar": "الغول"},
    "hint_32": {"ru": "Борис (кузнец)", "en": "Boris (blacksmith)", "fr": "Boris (le forgeron)",
                "uk": "Борис (коваль)", "pl": "Borys (kowal)", "ar": "بوريس (الحداد)"},
    # Настя: в озвучке он «директор ШКОЛЫ» (story_58 на всех языках),
    # а не Headmaster — так и подписываем.
    "hint_49": {"ru": "Директор школы", "en": "School Director", "fr": "Directeur de l'école",
                "uk": "Директор школи", "pl": "Dyrektor szkoły", "ar": "مدير المدرسة"},
    "hint_50": {"ru": "Директор школы", "en": "School Director", "fr": "Directeur de l'école",
                "uk": "Директор школи", "pl": "Dyrektor szkoły", "ar": "مدير المدرسة"},
    "hint_51": {"ru": "Директор школы", "en": "School Director", "fr": "Directeur de l'école",
                "uk": "Директор школи", "pl": "Dyrektor szkoły", "ar": "مدير المدرسة"},
    "hint_56": {"ru": "Директор школы", "en": "School Director", "fr": "Directeur de l'école",
                "uk": "Директор школи", "pl": "Dyrektor szkoły", "ar": "مدير المدرسة"},
    # Этапы на внешних ESP32
    "hint_15": {"ru": "Поезд", "en": "Train", "fr": "Train",
                "uk": "Поїзд", "pl": "Pociąg", "ar": "القطار"},
    "hint_16": {"ru": "Поезд", "en": "Train", "fr": "Train",
                "uk": "Поїзд", "pl": "Pociąg", "ar": "القطار"},
    "hint_9":  {"ru": "Волк", "en": "Wolf", "fr": "Loup",
                "uk": "Вовк", "pl": "Wilk", "ar": "الذئب"},
    "hint_7":  {"ru": "Чемоданы", "en": "Suitcases", "fr": "Valises",
                "uk": "Валізи", "pl": "Walizki", "ar": "الحقائب"},
    "hint_28": {"ru": "Сейф", "en": "Safe", "fr": "Coffre-fort",
                "uk": "Сейф", "pl": "Sejf", "ar": "الخزنة"},
}

# Соответствие «Playing Hint N» из логов ESP32 → id подсказки.
# В прошивке ESP32 знает только номер трека на своей SD-карте, а не id, поэтому
# связь восстановлена по числу подсказок у каждого устройства и по смыслу
# текстов: количество совпало у всех четырёх ровно.
# ПОРЯДОК внутри списка — единственное допущение этой таблицы. Если в чате
# у какого-то устройства текст поедет на одну подсказку, менять здесь.
ESP32_HINTS = {
    "train": ["hint_15_a", "hint_15_b", "hint_15_c", "hint_15_d", "hint_15_e",
              "hint_16_c", "hint_16_d", "hint_16_z"],
    "wolf": ["hint_9_a", "hint_9_b", "hint_9_c", "hint_9_d", "hint_9_e", "hint_9_f", "hint_9_z"],
    "chest": ["hint_7_a", "hint_7_b", "hint_7_c", "hint_7_d", "hint_7_z"],
    "safe": ["hint_28_a", "hint_28_b", "hint_28_c", "hint_28_z"],
}


def main():
    try:
        import openpyxl
    except ImportError:
        sys.exit("Нужен openpyxl: pip3 install openpyxl")

    tmp = "/tmp/clc_voice_export.xlsx"
    url = f"https://docs.google.com/spreadsheets/d/{SHEET}/export?format=xlsx"
    print("Тяну таблицу…")
    r = subprocess.run(["curl", "-sL", "--max-time", "90", "-o", tmp, url])
    if r.returncode != 0 or not os.path.exists(tmp):
        sys.exit("Не скачалось. Таблица открыта по ссылке на чтение — проверь интернет.")

    ws = openpyxl.load_workbook(tmp)["hints"]
    header = [str(c or "").strip().lower() for c in next(ws.iter_rows(min_row=1, max_row=1,
                                                                     values_only=True))]
    if "russian" not in header[1]:
        sys.exit(f"Шапка вкладки hints изменилась: {header[:4]} — проверь порядок столбцов")

    texts, latin_fix = {}, []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        hid = str(row[0]).strip()
        # В таблице встречается id с кириллической «с» (hint_51_с). На Pi файл
        # называется латиницей, и прошивка просит латиницей — поэтому чиним,
        # иначе текст к такой подсказке не найдётся никогда.
        fixed = hid.replace("с", "c").replace("а", "a").replace("е", "e")
        if fixed != hid:
            latin_fix.append(hid)
            hid = fixed
        texts[hid] = {lang: (row[i + 1] or "").strip()
                      for i, lang in enumerate(COLUMNS) if i + 1 < len(row)}

    data = {
        "_комментарий": ("Выгружено из Google-таблицы скриптом tools/export_hint_texts.py. "
                         "Руками не править: правится таблица, потом прогон скрипта."),
        # Ключи — префиксы и полные id, значения — имена по языкам. Сервер ищет
        # сначала точный id, потом префикс: так имя находится даже у реплики,
        # которой ещё нет в таблице.
        "speakers": SPEAKERS,
        "esp32": ESP32_HINTS,
        "texts": texts,
    }
    with open(OUT, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=1, sort_keys=True)

    print(f"Записано: {OUT}")
    print(f"  подсказок: {len(texts)}, размер: {os.path.getsize(OUT) // 1024} КБ")
    if latin_fix:
        print(f"  ⚠ id с кириллицей в таблице (починены при выгрузке): {latin_fix}")
    no_speaker = [h for h in texts
                  if h not in SPEAKERS and h.rsplit("_", 1)[0] not in SPEAKERS]
    if no_speaker:
        print(f"  ⚠ без персонажа (допиши в SPEAKERS): {no_speaker}")
    missing = [h for d in ESP32_HINTS.values() for h in d if h not in texts]
    if missing:
        print(f"  ⚠ в таблице нет текста для: {missing}")


if __name__ == "__main__":
    main()
